import streamlit as st
import pandas as pd
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- 1. НАСТРОЙКИ СТРАНИЦЫ ---
st.set_page_config(page_title="Аудит ИТ и ИБ 2026", layout="wide", page_icon="🛡️")

# --- 2. ЛОГОТИП И КОНТАКТЫ ---
if os.path.exists("logo.png"):
    st.image("logo.png", width=300)
else:
    st.title("Ivan Rudoy | IT Audit & Consulting")

st.markdown("### ")
st.divider()

st.title("📋 Опросник: Технический аудит ИТ и ИБ (2026)")

data = {}
score = 0

# --- БЛОК 1: ИНФРАСТРУКТУРА ---
st.header("Блок 1: Общая информация")

# 1.1 Конечные точки (АРМ)
st.subheader("1.1. Конечные точки (АРМ)")
total_arm = st.number_input("Общее количество АРМ (шт):", min_value=0, step=1)
data['1.1. Всего АРМ'] = total_arm

selected_os_arm = st.multiselect("Выберите ОС на АРМ:", ["Windows", "Linux", "macOS", "Другое"])
if selected_os_arm:
    for os_item in selected_os_arm:
        count_arm = st.number_input(f"Количество АРМ на {os_item}:", min_value=0, max_value=total_arm if total_arm > 0 else 1000000, step=1, key=f"arm_{os_item}")
        data[f"ОС АРМ ({os_item})"] = count_arm

st.write("---")

# 1.2 Серверная инфраструктура
st.subheader("1.2. Серверы")
col_s1, col_s2 = st.columns(2)
with col_s1:
    phys_servers = st.number_input("Количество физических серверов:", min_value=0, step=1)
    data['1.2. Физические серверы'] = phys_servers
with col_s2:
    virt_servers = st.number_input("Количество виртуальных серверов:", min_value=0, step=1)
    data['1.2. Виртуальные серверы'] = virt_servers

selected_os_srv = st.multiselect("Выберите ОС серверов:", ["Windows Server", "Linux", "Unix", "Другое"])
if selected_os_srv:
    for os_s in selected_os_srv:
        count_srv = st.number_input(f"Количество серверов на {os_s}:", min_value=0, step=1, key=f"srv_{os_s}")
        data[f"ОС Сервера ({os_s})"] = count_srv

st.write("---")

# 1.3 Виртуализация и 1.4 Почта
col_v1, col_v2 = st.columns(2)
with col_v1:
    st.subheader("1.3. Виртуализация")
    data['1.3. Виртуализация'] = st.multiselect("Системы виртуализации:", ["VMware", "Hyper-V", "Proxmox", "KVM", "Другое", "Нет"], key="virt_13")
with col_v2:
    st.subheader("1.4. Почтовая система")
    data['1.4. Почта'] = st.selectbox("Тип почты:", ["Exchange (On-Prem)", "Microsoft 365", "Google Workspace", "Yandex/Mail.ru Cloud", "Собственный сервер", "Нет"], key="mail_14")

st.write("---")

# 1.5 Информационные системы
st.subheader("1.5. Внутренние Информационные системы")
has_is = st.checkbox("Есть ли внутренние Информационные системы (1C, ERP, CRM)?", key="is_15")
if has_is:
    data['1.5. Внутренние ИС'] = st.text_input("Перечислите их через запятую:", key="is_list")
else:
    data['1.5. Внутренние ИС'] = "Нет"

# 1.6 Мониторинг
st.subheader("1.6. Система мониторинга")
has_mon = st.checkbox("Есть ли система мониторинга?", key="mon_16")
if has_mon:
    data['1.6. Мониторинг'] = st.selectbox("Выберите систему:", ["Zabbix", "Nagios", "PRTG", "Prometheus", "Другое"], key="mon_choice")
else:
    data['1.6. Мониторинг'] = "Нет"

# --- БЛОК 2: СЕТЬ ---
st.header("Блок 2: Сетевая инфраструктура и Интернет")
if st.toggle("Показать Блок 2"):
    net_types = ["Оптика", "Радиорелейная", "Спутник", "4G/5G", "Starlink"]
    
    c_n1, c_n2 = st.columns(2)
    with c_n1:
        data['2.1. Основной канал'] = st.selectbox("Тип канала (Основной):", net_types)
        data['2.1. Скорость осн. (mbit/s)'] = st.number_input("Скорость основного канала (mbit/s):", min_value=0)
    with c_n2:
        data['2.1. Резервный канал'] = st.selectbox("Тип канала (Резервный):", ["Нет"] + net_types)
        data['2.1. Скорость рез. (mbit/s)'] = st.number_input("Скорость резервного канала (mbit/s):", min_value=0)

    data['2.4. NGFW'] = st.text_input("Вендор Межсетевого экрана (NGFW):")
    if data['2.4. NGFW']: score += 20

    has_wifi = st.checkbox("Используется Wi-Fi?")
    if has_wifi:
        if st.checkbox("Есть ли Wi-Fi контроллер?"):
            data['2.5. Контроллер'] = st.text_input("Модель контроллера:")
        else:
            data['2.5. Контроллер'] = "Без контроллера"
        data['2.5. Число точек'] = st.number_input("Кол-во точек доступа:", min_value=0)

# --- БЛОК 3: ИБ ---
st.header("Блок 3: Информационная Безопасность")
st.info("Отметьте системы и укажите вендора (например: Kaspersky, InfoWatch, Fortinet)")

ib_list = {
    "DLP (Защита от утечек)": 15,
    "PAM (Контроль доступа)": 10,
    "SIEM/SOC (Мониторинг ИБ)": 20,
    "WAF (Защита Web)": 10,
    "EDR/Antimalware": 15,
    "Резервное копирование": 20
}

for label, pts in ib_list.items():
    col_ib1, col_ib2 = st.columns([1, 2])
    with col_ib1:
        is_on = st.checkbox(label, key=f"ib_chk_{label}")
    if is_on:
        with col_ib2:
            vendor = st.text_input(f"Вендор {label}:", key=f"ib_v_{label}")
            data[label] = f"Да ({vendor if vendor else 'не указан'})"
            score += pts
    else:
        data[label] = "Нет"

st.write("---")
if st.checkbox("3.7. Другое (дополнительные системы защиты)"):
    data['3.7. Прочие системы ИБ'] = st.text_area("Перечислите все, что мы не учли:")
else:
    data['3.7. Прочие системы ИБ'] = "Нет"

# --- ЭКСЕЛЬ ГЕНЕРАЦИЯ ---
def make_excel(results, final_score):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты аудита"

    # Стилизация
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Заголовок
    ws.merge_cells('A1:C1')
    ws['A1'] = "ТЕХНИЧЕСКИЙ АУДИТ 2026 - ЭКСПЕРТНЫЙ ОТЧЕТ"
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].font = Font(bold=True, size=12)

    # Индекс зрелости
    ws['A3'] = "ИНДЕКС ЗРЕЛОСТИ:"
    ws['B3'] = f"{final_score} / 100"
    color_code = "00B050" if final_score > 70 else "FFCC00" if final_score > 40 else "FF0000"
    ws['B3'].fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")

    # Таблица
    headers = ["Параметр", "Значение", "Анализ"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=i, value=h)
        cell.fill = header_fill
        cell.font = white_font

    for idx, (k, v) in enumerate(results.items(), 6):
        ws.cell(row=idx, column=1, value=k).border = thin_border
        ws.cell(row=idx, column=2, value=str(v)).border = thin_border
        
        # Авто-аналитика
        analysis = "В норме"
        if "Нет" in str(v) or v == 0:
            analysis = "РИСК / ТРЕБУЕТ ВНИМАНИЯ"
            ws.cell(row=idx, column=3).font = Font(color="FF0000", bold=True)
        ws.cell(row=idx, column=3, value=analysis).border = thin_border

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    wb.save(output)
    return output.getvalue()

# --- ФИНАЛ ---
st.divider()
if st.button("📊 Сформировать экспертный отчет"):
    if not data:
        st.error("Данные не заполнены!")
    else:
        f_score = min(score, 100)
        xlsx = make_excel(data, f_score)
        st.success(f"Готово! Индекс тех. зрелости: {f_score}/100")
        st.download_button("📥 Скачать Excel Аудит 2026", xlsx, "Audit_Expert_Report.xlsx")

st.info("Разработано Ivan Rudoy. По вопросам системной интеграции — звоните!")
